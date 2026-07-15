"""
Calibrate Omega^H, Omega^F, beta^H (and alpha, export shares, Domar weights)
for the 3-sector SOE model from Korea's official national input-output table.

DATA SOURCE
-----------
Bank of Korea, "2022 Input-Output Tables (Producer's price, Large-Sized, 33-sector)".
Downloaded 2026-07-15 from https://www.bok.or.kr/eng/bbs/E0000634/view.do?nttId=10088182
File: korea_raw/korea_io_2022.xlsx, sheets used:
  - 'Transaction_domestic(producer)'  domestic intermediate use, product x activity, 33x33
  - 'Transaction_imported(producer)'  imported intermediate use, product x activity, 33x33
  - 'Transaction(producer)'           combined table, used for value-added rows and gross
                                       output ('Total input', row label 9790)
All three sheets share the same 33-activity layout (row 5 = activity code, columns 3..35),
and the same row layout (rows 7..39 = the 33 commodities/activities in the same order).

SECTOR CONCORDANCE (33 BOK activities -> our 3 model sectors)
---------------------------------------------------------------
  Resource      = {A Agricultural/forestry/fishery, B Mining and quarrying}
  Manufacturing = {C01-C14 manufacturing subsectors, D Electricity/gas/steam,
                    E Water/sewage/waste, F Construction}
  Services      = {G..T: trade, transport, food/accommodation, communications, finance,
                    real estate, professional/business services, public admin, education,
                    health, arts, other services, "Others"}

Same concordance logic as the Chile calibration (build_chile_calibration.py): Resource =
upstream/tradable primary sector, Manufacturing = mid-chain industrial production (+
utilities/construction), Services = downstream, largely non-tradable, consumption-heavy.
Korea is the "developed, diversified manufacturing exporter" contrast to Chile's
commodity-exporter calibration.

MODEL CONVENTION
----------------
Same as build_chile_calibration.py: Y_it = A_it L_it^alpha_i * prod_j X_ijt^Omega^H_ij *
M_it^Omega^F_i, i.e. Omega^H_ij is sector i's (the BUYER's) cost share spent on domestic
sector j's output. The raw IO sheets are Product (row) x Activity (column), i.e. entry
[j, i] = flow FROM supplying sector j TO using activity i, so Omega^H_ij (buyer i, seller
j) = raw[j, i] / Y_i, the TRANSPOSE of the raw sheet.

OUTPUT
------
Writes korea_calibration_results.json with the full aggregated Omega^H (3x3), Omega^F,
alpha, beta^H, export shares, Domar weights and DC weights, in the same format as
chile_calibration_results.json.
"""

import json
from pathlib import Path

import numpy as np
import openpyxl

HERE = Path(__file__).parent
RAW = HERE / "korea_raw" / "korea_io_2022.xlsx"

# 33 BOK activity codes in sheet order (columns 3..35 / rows 7..39)
ACTIVITY_CODES = [
    "A", "B",
    "C01", "C02", "C03", "C04", "C05", "C06", "C07", "C08", "C09",
    "C10", "C11", "C12", "C13", "C14",
    "D", "E", "F",
    "G", "H", "I", "J", "K", "L", "M", "N", "O", "P", "Q", "R", "S", "T",
]
N33 = len(ACTIVITY_CODES)
assert N33 == 33

CONCORDANCE = {}
for code in ["A", "B"]:
    CONCORDANCE[code] = 0  # Resource
for code in ["C01", "C02", "C03", "C04", "C05", "C06", "C07", "C08", "C09",
             "C10", "C11", "C12", "C13", "C14", "D", "E", "F"]:
    CONCORDANCE[code] = 1  # Manufacturing
for code in ["G", "H", "I", "J", "K", "L", "M", "N", "O", "P", "Q", "R", "S", "T"]:
    CONCORDANCE[code] = 2  # Services

MACRO_NAMES = ["Resource", "Manufacturing", "Services"]
N3 = 3

ROW_START = 7  # first activity row in every sheet
COL_START = 3   # first activity column in every sheet


def read_flow_matrix(ws):
    """Read the 33x33 [product_row, activity_col] flow matrix."""
    mat = np.zeros((N33, N33))
    for i in range(N33):
        row = ROW_START + i
        for j in range(N33):
            col = COL_START + j
            v = ws.cell(row=row, column=col).value
            mat[i, j] = float(v) if v is not None else 0.0
    return mat


def read_final_demand_column(ws, col):
    vec = np.zeros(N33)
    for i in range(N33):
        row = ROW_START + i
        v = ws.cell(row=row, column=col).value
        vec[i] = float(v) if v is not None else 0.0
    return vec


def read_bottom_row(ws, row):
    """Read a bottom summary row (value added, total input, ...) across the 33 activity columns."""
    vec = np.zeros(N33)
    for j in range(N33):
        col = COL_START + j
        v = ws.cell(row=row, column=col).value
        vec[j] = float(v) if v is not None else 0.0
    return vec


def aggregate_matrix(mat33):
    agg = np.zeros((N3, N3))
    for i, ci in enumerate(ACTIVITY_CODES):
        for j, cj in enumerate(ACTIVITY_CODES):
            agg[CONCORDANCE[ci], CONCORDANCE[cj]] += mat33[i, j]
    return agg


def aggregate_vector(vec33):
    agg = np.zeros(N3)
    for i, ci in enumerate(ACTIVITY_CODES):
        agg[CONCORDANCE[ci]] += vec33[i]
    return agg


def main():
    wb = openpyxl.load_workbook(RAW, data_only=True)
    dom_ws = wb["Transaction_domestic(producer)"]
    imp_ws = wb["Transaction_imported(producer)"]
    comb_ws = wb["Transaction(producer)"]

    dom_use = read_flow_matrix(dom_ws)     # domestic intermediate use, product x activity
    imp_use = read_flow_matrix(imp_ws)     # imported intermediate use, product x activity

    hh_cons_col = 37   # 9111 Private final consumption expenditure
    exports_col = 43   # 9140 Exports
    hh_cons_33 = read_final_demand_column(comb_ws, hh_cons_col)
    exports_33 = read_final_demand_column(comb_ws, exports_col)

    # locate value-added / gross-output rows in the combined sheet by label in column A
    row_labels = {comb_ws.cell(row=r, column=1).value: r for r in range(38, 50)
                  if comb_ws.cell(row=r, column=1).value}
    va_row = row_labels["9690"]     # Total value added
    gross_row = row_labels["9790"]  # Total input (= gross output)

    VA_33 = read_bottom_row(comb_ws, va_row)
    Y_33 = read_bottom_row(comb_ws, gross_row)

    # --- aggregate to 3 macro-sectors ---
    dom_use_3 = aggregate_matrix(dom_use)
    imp_use_3 = aggregate_matrix(imp_use)
    Y3 = aggregate_vector(Y_33)
    VA3 = aggregate_vector(VA_33)
    hh_cons_3 = aggregate_vector(hh_cons_33)
    exports_3 = aggregate_vector(exports_33)

    # Omega^H[i,j] = buyer i's cost share spent on domestic sector j
    #              = dom_use_3[j, i] / Y3[i]   (transpose: raw is [seller_row, buyer_col])
    OmegaH = np.zeros((N3, N3))
    OmegaF = np.zeros(N3)
    alpha = np.zeros(N3)
    for i in range(N3):
        for j in range(N3):
            OmegaH[i, j] = dom_use_3[j, i] / Y3[i]
        OmegaF[i] = imp_use_3[:, i].sum() / Y3[i]
        alpha[i] = VA3[i] / Y3[i]

    # Cost shares should sum to ~1; renormalize to absorb margins/taxes wedge (basic- vs
    # purchaser-price gap), same normalization step as the Chile calibration.
    raw_sum = alpha + OmegaH.sum(axis=1) + OmegaF
    OmegaH = OmegaH / raw_sum[:, None]
    OmegaF = OmegaF / raw_sum
    alpha = alpha / raw_sum

    betaH = hh_cons_3 / hh_cons_3.sum()
    export_share = exports_3 / Y3

    # --- network objects, same formulas as the Chile calibration / presentation ---
    I3 = np.eye(N3)
    leontief_inv = np.linalg.inv(I3 - OmegaH)
    domar = betaH @ leontief_inv
    import_centrality = leontief_inv @ OmegaF

    # Raw Calvo reset probabilities: NOT identifiable from IO data. No
    # country-specific sector-level price-microdata estimate was found for
    # Korea, so this uses the SAME literature-sourced default as the
    # open_economy_network_{chile,korea,czechia}.mod files (kept in sync by
    # hand -- see the DELTA1-3 comment block there for the full derivation):
    # euro-area monthly price-change frequencies (Dhyne et al. 2005/ECB IPN)
    # converted to a quarterly Calvo reset probability via
    # delta_q = 1-(1-f_monthly)^3, cross-checked against Nakamura & Steinsson
    # (2008) US PPI durations for the Manufacturing sector.
    BETA = 0.99
    delta = np.array([0.90, 0.31, 0.16])  # Resource, Manufacturing, Services (see .mod file for sourcing)
    dhat = delta * (1 - BETA * (1 - delta)) / (1 - BETA * delta * (1 - delta))
    dc_weight_raw = domar * (1 - dhat) / dhat
    dc_weight = dc_weight_raw / dc_weight_raw.sum()

    results = {
        "source": "Bank of Korea, 2022 Input-Output Tables (Producer's price, Large-Sized, 33-sector)",
        "sectors": MACRO_NAMES,
        "gross_output_mnKRW2022": dict(zip(MACRO_NAMES, Y3.round(1))),
        "OmegaH": [[round(x, 4) for x in row] for row in OmegaH],
        "OmegaF": [round(x, 4) for x in OmegaF],
        "alpha": [round(x, 4) for x in alpha],
        "betaH": [round(x, 4) for x in betaH],
        "export_share_of_own_output": [round(x, 4) for x in export_share],
        "domar_weight": [round(x, 4) for x in domar],
        "import_centrality": [round(x, 4) for x in import_centrality],
        "dc_weight_literature_delta": {
            "note": "delta_i (Calvo reset prob.) not identifiable from IO data; literature-sourced default "
                    "(Dhyne et al. 2005 / ECB IPN, cross-checked vs. Nakamura & Steinsson 2008), "
                    "same value used in the .mod file -- see its DELTA1-3 comment for the full derivation.",
            "delta_raw_used": [round(x, 4) for x in delta],
            "dhat_derived": [round(x, 4) for x in dhat],
            "dc_weight": [round(x, 4) for x in dc_weight],
        },
    }

    out_path = HERE / "korea_calibration_results.json"
    out_path.write_text(json.dumps(results, indent=2))

    print(f"{'':16s}{'Resource':>12s}{'Manuf.':>12s}{'Services':>12s}")
    print(f"{'Omega^F':16s}" + "".join(f"{v:12.4f}" for v in OmegaF))
    print(f"{'beta^H':16s}" + "".join(f"{v:12.4f}" for v in betaH))
    print(f"{'Export sh.':16s}" + "".join(f"{v:12.4f}" for v in export_share))
    print(f"{'alpha':16s}" + "".join(f"{v:12.4f}" for v in alpha))
    print(f"{'Domar':16s}" + "".join(f"{v:12.4f}" for v in domar))
    print(f"{'Import cent.':16s}" + "".join(f"{v:12.4f}" for v in import_centrality))
    print()
    print("Full domestic IO matrix Omega^H[buyer i, seller j] (dense):")
    print(f"{'':16s}" + "".join(f"{n:>12s}" for n in MACRO_NAMES))
    for i, name in enumerate(MACRO_NAMES):
        print(f"{name:16s}" + "".join(f"{OmegaH[i, j]:12.4f}" for j in range(N3)))
    print(f"\nWrote {out_path}")


if __name__ == "__main__":
    main()
