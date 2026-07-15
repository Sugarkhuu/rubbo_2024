"""
Calibrate Omega^H, Omega^F, beta^H (and alpha, export shares, Domar weights)
for the 3-sector SOE model from Chile's official national input-output table.

DATA SOURCE
-----------
Banco Central de Chile, "Cuentas Nacionales de Chile - Matriz Insumo Producto".
Compilation reference: CdeR 2018. Reference year of the tables: 2023.
Page:     https://www.bcentral.cl/web/banco-central/areas/estadisticas/matriz-insumo-producto/cuadros-mip-excel
Files used (downloaded 2026-07-15, values in billions of 2023 CLP):
  - mip_12x12.xlsx  "MIP 12x12"      -> domestic direct-coefficient matrix (sheet '2')
  - cou_12x12.xlsx  "Cuadros 12x12"  -> Supply-Use tables (25 sheets), of which we use:
        sheet 19  Utilizacion intermedia NACIONAL, precios basico   (domestic intermediate use)
        sheet 21  Utilizacion intermedia IMPORTADA, precios basico  (imported intermediate use)
        sheet 20  Utilizacion final NACIONAL, precios basico        (household consumption, exports of
                                                                      domestically produced goods)
        sheet 23  Cuadrante de valor agregado                       (intermediate consumption & value
                                                                      added by activity -> gross output)

Both tables use the Central Bank's 12-activity classification (see SECTOR_NAMES below).

SECTOR CONCORDANCE (12 CdeR activities -> our 3 model sectors)
---------------------------------------------------------------
  Resource      = {1 Agropecuario-silvicola y Pesca, 2 Mineria}
  Manufacturing = {3 Industria manufacturera, 4 Electricidad/gas/agua/gestion de desechos, 5 Construccion}
  Services      = {6 Comercio-hoteles-restaurantes, 7 Transporte/comunicaciones, 8 Intermediacion
                   financiera, 9 Servicios inmobiliarios, 10 Servicios empresariales,
                   11 Servicios personales, 12 Administracion publica}

This mirrors the paper's story: Resource = upstream/tradable primary sector (mining is Chile's
copper-export engine, the real-world analogue of an "oil exporter"), Manufacturing = mid-chain
industrial production, Services = downstream, largely non-tradable, consumption-heavy sector.

MODEL CONVENTION
----------------
The model writes  Y_it = A_it L_it^alpha_i * prod_j X_ijt^Omega^H_ij * M_it^Omega^F_i,
i.e. Omega^H_ij is sector i's (the BUYER's) cost share spent on domestic sector j's output,
and alpha_i + sum_j Omega^H_ij + Omega^F_i = 1 (three cost shares: labor, domestic inputs,
imports -- there is no separate capital factor, so "alpha_i" here is read as the *total*
value-added share, i.e. labor + capital + net taxes together, since the model has no
explicit capital input).

The raw IO sheets are Product (row) x Activity (column), i.e. entry [j, i] = flow FROM
supplying sector j TO using activity i. So Omega^H_ij (buyer i, seller j) = raw[j, i] / Y_i,
which is the TRANSPOSE of the raw sheet.

OUTPUT
------
Prints an old-vs-new comparison table and writes chile_calibration_results.json with the
full aggregated Omega^H (3x3), Omega^F, alpha, beta^H, export shares, Domar weights and
DC weights (using the same formulas as the presentation's Network Properties slide).
"""

import json
from pathlib import Path

import numpy as np
import openpyxl

HERE = Path(__file__).parent

SECTOR_NAMES = {
    1: "Agropecuario-silvicola y Pesca",
    2: "Mineria",
    3: "Industria manufacturera",
    4: "Electricidad, gas, agua y gestion de desechos",
    5: "Construccion",
    6: "Comercio, hoteles y restaurantes",
    7: "Transporte, comunicaciones y servicios de informacion",
    8: "Intermediacion financiera",
    9: "Servicios inmobiliarios y de vivienda",
    10: "Servicios empresariales",
    11: "Servicios personales",
    12: "Administracion publica",
}

# 12 CdeR activities -> {0: Resource, 1: Manufacturing, 2: Services}
CONCORDANCE = {
    1: 0, 2: 0,
    3: 1, 4: 1, 5: 1,
    6: 2, 7: 2, 8: 2, 9: 2, 10: 2, 11: 2, 12: 2,
}
MACRO_NAMES = ["Resource", "Manufacturing", "Services"]
N12, N3 = 12, 3


def read_matrix_sheet(wb, sheet_name):
    """Read a 12x12 'Producto x Actividad' matrix from a COU sheet (rows/cols 1..12)."""
    ws = wb[sheet_name]
    # header row holds column activity codes 1..12 in columns C..N (col index 3..14);
    # data rows hold the product code in column C (index 3) and values in D..O (4..15)
    header_row = None
    for r in range(1, 15):
        vals = [ws.cell(row=r, column=c).value for c in range(3, 15)]
        if vals == list(range(1, 13)):
            header_row = r
            break
    if header_row is None:
        raise ValueError(f"Could not locate 1..12 header row in sheet {sheet_name}")

    data_start = None
    for r in range(header_row + 1, header_row + 6):
        if ws.cell(row=r, column=2).value == 1:
            data_start = r
            break
    if data_start is None:
        raise ValueError(f"Could not locate data start row in sheet {sheet_name}")

    mat = np.zeros((N12, N12))
    for i in range(N12):
        row = data_start + i
        product_code = ws.cell(row=row, column=2).value
        assert product_code == i + 1, (sheet_name, row, product_code)
        for j in range(N12):
            v = ws.cell(row=row, column=3 + j).value
            mat[i, j] = float(v) if v is not None else 0.0
    return mat  # mat[product_row, activity_col] = flow from product (row) to activity (col)


def read_final_use_sheet(wb, sheet_name):
    """Read a 'Utilizacion final ...' sheet: returns dict of column-name -> length-12 vector,
    indexed by product 1..12."""
    ws = wb[sheet_name]
    header_row = None
    for r in range(1, 15):
        if ws.cell(row=r, column=4).value == "Consumo intermedio":
            header_row = r
            break
    if header_row is None:
        raise ValueError(f"Could not find header row in {sheet_name}")

    col_names = []
    for c in range(4, 12):
        v = ws.cell(row=header_row, column=c).value
        col_names.append(v)

    data_start = None
    for r in range(header_row + 1, header_row + 6):
        if ws.cell(row=r, column=2).value == 1:
            data_start = r
            break
    if data_start is None:
        raise ValueError(f"Could not locate data start row in sheet {sheet_name}")

    out = {name: np.zeros(N12) for name in col_names if name}
    for i in range(N12):
        row = data_start + i
        product_code = ws.cell(row=row, column=2).value
        assert product_code == i + 1, (sheet_name, row, product_code)
        for c, name in zip(range(4, 12), col_names):
            if not name:
                continue
            v = ws.cell(row=row, column=c).value
            out[name][i] = float(v) if v is not None else 0.0
    return out


def read_value_added_sheet(wb, sheet_name="23"):
    """Returns (intermediate_consumption[12], value_added[12]) by ACTIVITY (not product)."""
    ws = wb[sheet_name]
    header_row = None
    for r in range(1, 15):
        vals = [ws.cell(row=r, column=c).value for c in range(4, 16)]
        if vals == list(range(1, 13)):
            header_row = r
            break
    if header_row is None:
        raise ValueError("Could not find activity header row in value-added sheet")

    def row_values(label):
        for r in range(header_row + 1, header_row + 6):
            if ws.cell(row=r, column=2).value == label:
                return np.array([float(ws.cell(row=r, column=4 + j).value or 0.0) for j in range(N12)])
        raise ValueError(f"Row '{label}' not found")

    ci = row_values("Consumo intermedio")
    va = row_values("Valor agregado")
    return ci, va


def aggregate_matrix(mat12):
    """Aggregate a 12x12 [product_row, activity_col] flow matrix to 3x3 macro-sector flows."""
    agg = np.zeros((N3, N3))
    for i in range(N12):
        for j in range(N12):
            agg[CONCORDANCE[i + 1], CONCORDANCE[j + 1]] += mat12[i, j]
    return agg


def aggregate_vector(vec12):
    agg = np.zeros(N3)
    for i in range(N12):
        agg[CONCORDANCE[i + 1]] += vec12[i]
    return agg


def main():
    mip = openpyxl.load_workbook(HERE / "mip_12x12.xlsx", data_only=True)
    cou = openpyxl.load_workbook(HERE / "cou_12x12.xlsx", data_only=True)

    dom_use = read_matrix_sheet(cou, "19")   # domestic intermediate use, product x activity
    imp_use = read_matrix_sheet(cou, "21")   # imported intermediate use, product x activity
    final_nat = read_final_use_sheet(cou, "20")  # national final use (household C, exports, ...)
    ci_by_activity, va_by_activity = read_value_added_sheet(cou, "23")

    gross_output_12 = ci_by_activity + va_by_activity  # Y_i, 12 activities

    # --- aggregate to 3 macro-sectors ---
    dom_use_3 = aggregate_matrix(dom_use)     # [product_row, activity_col], 3x3
    imp_use_3 = aggregate_matrix(imp_use)     # [product_row, activity_col], 3x3
    Y3 = aggregate_vector(gross_output_12)    # gross output by macro activity
    VA3 = aggregate_vector(va_by_activity)
    hh_cons_3 = aggregate_vector(final_nat["Consumo de hogares"])
    exports_3 = aggregate_vector(final_nat["Exportaciones"])

    # Omega^H[i,j] = buyer i's cost share spent on domestic sector j
    #              = dom_use_3[j, i] / Y3[i]   (transpose: raw is [seller_row, buyer_col])
    OmegaH = np.zeros((N3, N3))
    OmegaF = np.zeros(N3)
    alpha = np.zeros(N3)
    for i in range(N3):
        for j in range(N3):
            OmegaH[i, j] = dom_use_3[j, i] / Y3[i]
        OmegaF[i] = imp_use_3[:, i].sum() / Y3[i]
        alpha[i] = VA3[i] / Y3[i]

    # Cost shares should sum to ~1; renormalize to absorb margins/taxes wedge (basic- vs
    # purchaser-price gap in the raw sheets), a standard step when mapping SUT data to a
    # Cobb-Douglas cost identity.
    raw_sum = alpha + OmegaH.sum(axis=1) + OmegaF
    OmegaH = OmegaH / raw_sum[:, None]
    OmegaF = OmegaF / raw_sum
    alpha = alpha / raw_sum

    betaH = hh_cons_3 / hh_cons_3.sum()
    export_share = exports_3 / Y3  # share of sector's OWN output that is exported

    # --- network objects, same formulas as the presentation ---
    I3 = np.eye(N3)
    leontief_inv = np.linalg.inv(I3 - OmegaH)          # (I - Omega^H)^{-1}
    domar = betaH @ leontief_inv                        # lambda_D,i
    import_centrality = leontief_inv @ OmegaF           # M_i = [(I-OmegaH)^-1 Omega^F 1]_i

    delta_hat = np.array([0.90, 0.50, 0.10])  # placeholder derived Calvo slopes (delta_i not in IO data)
    dc_weight_raw = domar * (1 - delta_hat) / delta_hat
    dc_weight = dc_weight_raw / dc_weight_raw.sum()

    results = {
        "source": "Banco Central de Chile, Cuadros 12x12, CdeR 2018, ano de referencia 2023",
        "sectors": MACRO_NAMES,
        "gross_output_bnCLP2023": dict(zip(MACRO_NAMES, Y3.round(1))),
        "OmegaH": [[round(x, 4) for x in row] for row in OmegaH],
        "OmegaF": [round(x, 4) for x in OmegaF],
        "alpha": [round(x, 4) for x in alpha],
        "betaH": [round(x, 4) for x in betaH],
        "export_share_of_own_output": [round(x, 4) for x in export_share],
        "domar_weight": [round(x, 4) for x in domar],
        "import_centrality": [round(x, 4) for x in import_centrality],
        "dc_weight_placeholder_delta": {
            "note": "delta_i not identifiable from IO data; placeholder Calvo slopes used, replace with estimated stickiness",
            "delta_hat_used": list(delta_hat),
            "dc_weight": [round(x, 4) for x in dc_weight],
        },
    }

    out_path = HERE / "chile_calibration_results.json"
    out_path.write_text(json.dumps(results, indent=2))

    # --- print old (invented) vs new (data-based) comparison ---
    old = {
        "OmegaH_21": 0.20, "OmegaH_32": 0.25,
        "OmegaF": [0.30, 0.10, 0.05],
        "betaH": [0.05, 0.15, 0.80],
        "export_share": [0.65, 0.20, 0.00],
    }
    print("=" * 70)
    print("OLD (invented) vs NEW (Chile IO 2023-vintage, CdeR2018) calibration")
    print("=" * 70)
    print(f"{'':16s}{'Resource':>12s}{'Manuf.':>12s}{'Services':>12s}")
    print(f"{'Omega^F (new)':16s}" + "".join(f"{v:12.4f}" for v in OmegaF))
    print(f"{'Omega^F (old)':16s}" + "".join(f"{v:12.2f}" for v in old['OmegaF']))
    print(f"{'beta^H (new)':16s}" + "".join(f"{v:12.4f}" for v in betaH))
    print(f"{'beta^H (old)':16s}" + "".join(f"{v:12.2f}" for v in old['betaH']))
    print(f"{'Export sh.(new)':16s}" + "".join(f"{v:12.4f}" for v in export_share))
    print(f"{'Export sh.(old)':16s}" + "".join(f"{v:12.2f}" for v in old['export_share']))
    print(f"{'alpha (new)':16s}" + "".join(f"{v:12.4f}" for v in alpha))
    print(f"{'Domar (new)':16s}" + "".join(f"{v:12.4f}" for v in domar))
    print(f"{'Import cent.new':16s}" + "".join(f"{v:12.4f}" for v in import_centrality))
    print()
    print("Full domestic IO matrix Omega^H[buyer i, seller j] (new, dense, NOT triangular):")
    print(f"{'':16s}" + "".join(f"{n:>12s}" for n in MACRO_NAMES))
    for i, name in enumerate(MACRO_NAMES):
        print(f"{name:16s}" + "".join(f"{OmegaH[i, j]:12.4f}" for j in range(N3)))
    print(f"\n(old model only had Omega^H_21={old['OmegaH_21']}, Omega^H_32={old['OmegaH_32']}, "
          f"rest zero -- i.e. a pure triangular chain)")
    print(f"\nWrote {out_path}")


if __name__ == "__main__":
    main()
